import os
import logging
from time import sleep

from PyQt5 import QtCore
from PyQt5.QtCore import QObject
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill

logger = logging.getLogger('Client_UI.components.statistic_management.statistic')


class StatisticManager(QObject):
    """Формирование статистики и запись в файл xlsx"""
    diag_signal = QtCore.pyqtSignal(str)
    finish_signal = QtCore.pyqtSignal()

    columns = ("№  ", "Команды", "Вид спорта", "Дата", "ТМ/Коэфф.", "ТБ/Коэфф.", "Скриншот", "Ссылка", "Ставка",
               "Баланс до", "Баланс после", "Время размещения", "Сумма голов", "Результат", "Доход", "Букмекер",
               "Букмекер", "Текущий баланс")
    font = 'Calibri'
    font_size = 12
    thins = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")
    imitation_start_balance = 10000

    def __init__(self, event_data: list = None, results: dict = None):
        super(StatisticManager, self).__init__()
        self.event_data = event_data
        self.results = results

    def create_or_open(self) -> Workbook | None:
        """Создание директории для файла статистики и файла статистики если отсутствуют"""
        if not os.path.exists(settings.STATS_DIR):
            os.mkdir(settings.STATS_DIR)

        if not os.path.exists(settings.STATS_FILE_NAME):
            wb = Workbook()
            ws_betting = wb.active
            ws_betting.title = "Ставки на деньги"
            for col in range(1, len(self.columns)+1):
                cell = ws_betting.cell(row=1, column=col, value=self.columns[col-1])
                cell.alignment = Alignment(horizontal="center")
                # выравнивание по ширине ячейки
                cell.font = Font(name=self.font, size=self.font_size)
                len_cell = len(str(cell.value))
                new_width_col = len_cell * self.font_size ** (self.font_size * 0.008)
                ws_betting.column_dimensions[cell.column_letter].width = new_width_col
                ws_betting.column_dimensions[cell.column_letter].alignment = Alignment(horizontal="center")

            ws_betting.column_dimensions['G'].alignment = Alignment(horizontal="left")
            ws_betting.column_dimensions['H'].alignment = Alignment(horizontal="left")
            ws_betting.column_dimensions['B'].width = 30
            ws_betting.column_dimensions['D'].width = 12
            ws_betting.column_dimensions['G'].width = 30
            ws_betting.column_dimensions['H'].width = 30

            # заливка таблицы текущего баланса
            for col in range(len(self.columns) - 1, len(self.columns) + 1):
                cell = ws_betting.cell(row=1, column=col)
                cell.fill = PatternFill(fgColor="fff794", fill_type="solid")
                cell.border = Border(top=self.double, bottom=self.double, left=self.double, right=self.double)

            ws_imitation = wb.copy_worksheet(ws_betting)
            ws_imitation.title = "Имитация ставок"
            return wb
        else:
            for _ in range(5):
                try:
                    wb = load_workbook(settings.STATS_FILE_NAME)
                    return wb
                except BaseException as ex:
                    message = 'Ошибка открытия файла статистики'
                    self.diag_signal.emit(message)
                    logger.info(f'{message}: {ex}')
                    if _ == 4:
                        message = 'Не удалось получить доступ к файлу статистики'
                        self.diag_signal.emit(message)
                        logger.info(f'{message}: {ex}')
                    sleep(0.5)

    def insert_data(self) -> None:
        """Заполнение данными документа xlsx"""
        wb = self.create_or_open()
        if not wb:
            self.finish_signal.emit()
            return
        try:
            if self.event_data[0]['bet_imitation']:
                ws = wb["Имитация ставок"]
            else:
                ws = wb["Ставки на деньги"]

            event_nums = []
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=len(ws['A'])):
                for cell in row:
                    if cell.value:
                        event_nums.append(int(cell.value))
            if event_nums:
                event_num = max(event_nums) + 1
            else:
                event_num = 1

            for data in self.event_data:
                empty_row_num = len(ws['A']) + 1
                ws.cell(row=empty_row_num, column=1).value = event_num
                ws.cell(row=empty_row_num, column=2).value = data['teams']
                ws.cell(row=empty_row_num, column=3).value = data['game_type']
                ws.cell(row=empty_row_num, column=4).value = data['date']
                if data['total_koeff_type'] == 'under':
                    ws.cell(row=empty_row_num, column=5).value = f"{data['total_nominal']}/{data['total_koeff']}"
                    ws.cell(row=empty_row_num, column=6).value = '-'
                else:
                    ws.cell(row=empty_row_num, column=5).value = '-'
                    ws.cell(row=empty_row_num, column=6).value = f"{data['total_nominal']}/{data['total_koeff']}"
                try:
                    screenshot_name = data['screenshot_name'].split('/')[-1]
                except BaseException:
                    screenshot_name = data['screenshot_name']
                ws.cell(row=empty_row_num, column=7).value = screenshot_name
                ws.cell(row=empty_row_num, column=8).hyperlink = data['url']
                ws.cell(row=empty_row_num, column=8).value = data['url']
                ws.cell(row=empty_row_num, column=9).value = data['bet_size']
                if data['bet_imitation'] and event_num == 1:
                    ws.cell(row=empty_row_num, column=10).value = self.imitation_start_balance
                    ws.cell(row=empty_row_num, column=11).value = self.imitation_start_balance - float(data['bet_size'])
                    for bkmkr_name_cell, cur_balance_cell in zip(ws['Q'], ws['R']):
                        if bkmkr_name_cell.value == data['bookmaker']:
                            cur_balance_cell.value = self.imitation_start_balance - float(data['bet_size'])
                            break
                        elif not bkmkr_name_cell.value:
                            bkmkr_name_cell.value = data['bookmaker']
                            cur_balance_cell.value = self.imitation_start_balance - float(data['bet_size'])
                            break

                elif data['bet_imitation'] and event_num > 1:
                    for bkmkr_name_cell, cur_balance_cell in zip(ws['Q'], ws['R']):
                        if bkmkr_name_cell.value == data['bookmaker']:
                            ws.cell(row=empty_row_num, column=10).value = cur_balance_cell.value
                            ws.cell(row=empty_row_num, column=11).value = cur_balance_cell.value - float(data['bet_size'])
                            cur_balance_cell.value = cur_balance_cell.value - float(data['bet_size'])
                            break
                        elif not bkmkr_name_cell.value:
                            ws.cell(row=empty_row_num, column=10).value = self.imitation_start_balance
                            ws.cell(row=empty_row_num, column=11).value = self.imitation_start_balance - float(data['bet_size'])
                            bkmkr_name_cell.value = data['bookmaker']
                            cur_balance_cell.value = self.imitation_start_balance - float(data['bet_size'])
                            break
                else:
                    ws.cell(row=empty_row_num, column=10).value = data['start_balance']
                    ws.cell(row=empty_row_num, column=11).value = data['balance_after_bet']
                    for bkmkr_name_cell, cur_balance_cell in zip(ws['Q'], ws['R']):
                        if bkmkr_name_cell.value == data['bookmaker']:
                            cur_balance_cell.value = data['balance_after_bet']
                            break
                        elif not bkmkr_name_cell.value:
                            bkmkr_name_cell.value = data['bookmaker']
                            cur_balance_cell.value = data['balance_after_bet']
                            break

                ws.cell(row=empty_row_num, column=12).value = data['betting_time']
                ws.cell(row=empty_row_num, column=15).value = f"-{data['bet_size']}"
                ws.cell(row=empty_row_num, column=16).value = data['bookmaker']

            # заливка разделительной линии
            empty_row_num = len(ws['A']) + 1
            for col in range(1, len(self.columns)-1):
                cell = ws.cell(row=empty_row_num, column=col)
                cell.fill = PatternFill(fgColor="000000", fill_type="solid")
                cell.border = Border(top=self.double, bottom=self.double, left=self.thins, right=self.thins)

            # заливка таблицы текущего баланса
            for col in range(len(self.columns) - 1, len(self.columns) + 1):
                for row in range(2, len(ws['Q'])):
                    cell = ws.cell(row=row, column=col)
                    if not cell.value:
                        continue
                    cell.fill = PatternFill(fgColor="fff794", fill_type="solid")
                    cell.border = Border(top=self.double, bottom=self.double, left=self.double, right=self.double)

            wb.save(settings.STATS_FILE_NAME)
            message = f'Данные по сделанной ставке успешно записаны в файл статистики {settings.STATS_FILE_NAME}'
            logger.info(message)
            TelegramService.send_xlsx(settings.STATS_FILE_NAME)
        except BaseException as ex:
            message = f'Ошибка при записи данных по сделанной ставке в файл статистики {settings.STATS_FILE_NAME}'
            logger.info(f'{message}: {ex}')
            TelegramService.send_text(message)

        self.diag_signal.emit(message)
        self.finish_signal.emit()

    def insert_results(self) -> None:
        """Внесение результатов событий на которые были сделаны ставки"""
        wb = self.create_or_open()
        if not wb:
            self.finish_signal.emit()
            return
        try:
            any_event_data = list(self.results.values())[0].get('event_data')
            if any_event_data[0]['bet_imitation']:
                ws = wb["Имитация ставок"]
            else:
                ws = wb["Ставки на деньги"]
        except BaseException as ex:
            logger.info(ex)
            self.finish_signal.emit()
            return
        try:
            for event_key, event_result in self.results.items():
                url_1 = event_key.replace('https:/', 'https://').split('$$')[0]
                url_2 = event_key.replace('https:/', 'https://').split('$$')[1]
                result = event_result.get('result')
                goals_sum = sum(list(map(int, result.split(':'))))

                for row in range(1, len(ws['A'])):
                    if (ws.cell(row=row, column=8).value == url_1 or ws.cell(row=row, column=8).value == url_2) \
                            and not ws.cell(row=row, column=14).value:
                        ws.cell(row=row, column=14).value = result
                        ws.cell(row=row, column=13).value = goals_sum
                        total_under_col_data = ws.cell(row=row, column=5).value
                        total_over_col_data = ws.cell(row=row, column=6).value
                        if total_under_col_data and total_under_col_data != '-':
                            total_nominal = float(total_under_col_data.replace(',', '.').split('/')[0])
                            total_coeff = float(total_under_col_data.replace(',', '.').split('/')[1])
                            if goals_sum < total_nominal:
                                ws.cell(row=row, column=15).value = float(ws.cell(row=row, column=15).value) + \
                                                                    total_coeff * float(ws.cell(row=row, column=9).value)
                                fgColor = "39CB05"
                            else:
                                fgColor = "fe0101"
                        else:
                            total_nominal = float(total_over_col_data.replace(',', '.').split('/')[0])
                            total_coeff = float(total_over_col_data.replace(',', '.').split('/')[1])
                            if goals_sum > total_nominal:
                                ws.cell(row=row, column=15).value = float(ws.cell(row=row, column=15).value) + \
                                                                    total_coeff * float(ws.cell(row=row, column=9).value)
                                fgColor = "39CB05"
                            else:
                                fgColor = "fe0101"

                        # заливка красным или зеленым строки в зависимости от итогового результата
                        for col in range(1, len(self.columns) - 1):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = PatternFill(fgColor=fgColor, fill_type="solid")

                        # подсчет итоговых балансов
                        for bkmkr_name_cell, cur_balance_cell in zip(ws['Q'], ws['R']):
                            if bkmkr_name_cell.value == ws.cell(row=row, column=16).value and float(ws.cell(row=row, column=15).value) > 0:
                                cur_balance_cell.value = cur_balance_cell.value + float(ws.cell(row=row, column=15).value)
                                break

            wb.save(settings.STATS_FILE_NAME)
            message = f'Результаты по сделанным ставкам успешно записаны в файл статистики {settings.STATS_FILE_NAME}'
            logger.info(message)
            TelegramService.send_xlsx(settings.STATS_FILE_NAME)
        except BaseException as ex:
            message = f'Ошибка при записи результатов событий по которым сделаны ставки в файл статистики {settings.STATS_FILE_NAME}'
            logger.info(f'{message}: {ex}')
            TelegramService.send_text(message)

        self.diag_signal.emit(message)
        self.finish_signal.emit()


if __name__ == '__main__':
    from pathlib import Path
    BASE_DIR = Path(__file__).resolve().parent.parent
    class settings:
        STATS_DIR = BASE_DIR / "statistic"
        STATS_FILE_NAME = STATS_DIR / "statistic.xlsx"
    event_data = [{'game_type': 'Soccer',
                   'balance_after_bet': 90,
                   'betting_time': '100',
                   'screenshot_name': 'D:/screenshots/leon.png',
                   'bet_imitation': True,
                   'start_balance': 100,
                   'total_koeff': 2.1,
                   'total_koeff_type': 'under',
                   'bet_size': 10,
                   'total_nominal': 2.5,
                   'url': 'link',
                   'teams': 'ЦСКА - Спартак',
                   'bookmaker': 'leon',
                   'date': '2025-01-12'},
                  {'game_type': 'Soccer',
                   'balance_after_bet': 110,
                   'betting_time': '80',
                   'screenshot_name': 'D:/screenshots/olimp.png',
                   'bet_imitation': True,
                   'start_balance': 120,
                   'total_koeff': 2.1,
                   'total_koeff_type': 'over',
                   'bet_size': 10,
                   'total_nominal': 2.5,
                   'url': 'link',
                   'teams': 'ЦСКА - Спартак',
                   'bookmaker': 'olimp',
                   'date': '2025-01-12'}]
    results = {}
    statistic = StatisticManager(event_data, results=results)
    statistic.insert_data()
    statistic.insert_results()
else:
    from .. import settings
    from ..telegram import TelegramService



# ("№  ", "Команды БК1", "Команды БК2", "ТМ/Коэфф. БК1", "ТБ/Коэфф. БК2", "ТБ/Коэфф. БК1", "ТМ/Коэфф. БК2",
#                "Коэффициент БК1", "Коэффициент БК2", "Скриншот БК1", "Скриншот БК2", "Ссылка БК1", "Ссылка БК2",
#                "Ставка БК1", "Результат БК1", "Банкрол БК1", "Ставка БК2", "Результат БК2", "Сумма голов", "Дата БК1",
#                "Дата БК2")