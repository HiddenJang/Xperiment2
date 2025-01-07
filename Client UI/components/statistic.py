import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill

from . import settings

class Statistic:
# функция записи в файл информации о сделанной имитационной ставке
    def chosen_event_data(event: list, shotname_after: str, element_takecoeff: list, bet_size: int): # получение данных о найденном событии из модуля GraphicalMODULE_1 функция tree_on_select
        l = 0 # флаг для проверки наличия в файле стартового банкрола 10000 в соответствующем столбце LEON
        x = 0 # флаг для проверки наличия в файле стартового банкрола 10000 в соответствующем столбце MELBET
        event_data = event
        if not os.path.exists(settings.STATS_DIR):
            os.mkdir(settings.STATS_DIR)
            row = 2 # номер ближайшей пустой строки
            wb = Workbook()  # создание  файла statistic.xlsx при помощи класса workbook
            ws_imitation = wb.create_sheet("Имитация ставок")
            ws = wb.active  # создание страницы в файле book
            ws.title = 'TMI' # название страницы в файле
            ws['A1'] = "№" #1 # названия столбцов
            ws['B1'] = "Команды первой БК" #2
            ws['C1'] = "Команды второй БК" #3
            ws['D1'] = "ТМ/Коэфф.первой БК" #4
            ws['E1'] = "ТБ/Коэфф.второй БК" #5
            ws['F1'] = "ТБ/Коэфф.первой БК" #6
            ws['G1'] = "ТМ/Коэфф.второй БК" #7
            ws['H1'] = "Коэффициент первой БК" #8
            ws['I1'] = "Коэффициент второй БК" #9
            ws['J1'] = "Название скриншота первой БК" #10
            ws['K1'] = "Название скриншота второй БК" #11
            ws['L1'] = "Ссылка на событие первой БК" #12
            ws['M1'] = "Ссылка на событие второй БК" #13
            ws['N1'] = "Размер ставки первой БК" #14
            ws['O1'] = "Результат первой БК" #15
            ws['P1'] = "Банкрол первой БК" #16
            ws['Q1'] = "Размер ставки второй БК" #17
            ws['R1'] = "Результат второй БК" #18
            ws['S1'] = "Банкрол второй БК" #19
            ws['T1'] = "Сумма голов"  # 20
            ws.column_dimensions['A'].width = 5 # ширина столбца
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 19
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 38
            ws.column_dimensions['K'].width = 38
            ws.column_dimensions['L'].width = 25
            ws.column_dimensions['M'].width = 25
            ws.column_dimensions['N'].width = 20
            ws.column_dimensions['O'].width = 16
            ws.column_dimensions['P'].width = 16
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 16
            ws.column_dimensions['S'].width = 16
            ws.column_dimensions['T'].width = 16
        else: # если файл создан уже, то открываем для записи
            n=1
            while True:
                try:
                    wb = load_workbook(f"{path}")
                    break
                except Exception as ex:
                    if n == 5:
                        logger.info(f'Ошибка открытия файла статистики. Статистика события {element_takecoeff[0]} не будет сформирована.')
                        return
                    n += 1
                    logger.info(f'Ошибка открытия файла статистики, событие {element_takecoeff[0]}, попытка: {n}. {ex}')
                    sleep(1)
            ws = wb["TMI"] # делаем страницу TMI активной
            row = 1
            while ws.cell(row=row, column=1).value != None: # определяем номер ближайшей пустой строки
                row+=1
            # стираем значения ИТОГО ранее занесенные в столбцы 15 и 18
            ws.cell(row=row, column=15).value = None
            ws.cell(row=row, column=18).value = None
        # проверки наличия в файле стартовых банкролов 10000 в соответствующих столбцах и присвоение последних значений банкролов
        for i in range(2, row+1):
            if ws.cell(row=i, column=16).value != None:
                bankrol_first = ws.cell(row=i, column=16).value
                l+=1
            if ws.cell(row=i, column=19).value != None:
                bankrol_second = ws.cell(row=i, column=19).value
                x+=1
        if l == 0:
            bankrol_first = '10000'
        if x == 0:
            bankrol_second = '10000'
        ws.cell(row=row, column=1, value=row) # внесение в первый столбец номера строки
        for i in range(0,len(event_data)-2):
            ws.cell(row=row, column=(i+2), value=event_data[i])  # обращаясь к листу записываем по координатам ячейки (row,column) значение (данные), 4 штуки - имя, цена, и т.д.
        if element_takecoeff[0] == 'LEON':
            if 'leon.ru' in event[6]:
                ws.cell(row=row, column=8, value=element_takecoeff[1])
                ws.cell(row=row, column=10, value=shotname_after)
                ws.cell(row=row, column=14, value=bet_size)
                ws.cell(row=row, column=16, value=(float(bankrol_first)-bet_size))
            else:
                ws.cell(row=row, column=9, value=element_takecoeff[1])
                ws.cell(row=row, column=11, value=shotname_after)
                ws.cell(row=row, column=17, value=bet_size)
                ws.cell(row=row, column=19, value=(float(bankrol_second) - bet_size))
        elif element_takecoeff[0] == 'MELBET':
            if 'melbet.ru' in event[6]:
                ws.cell(row=row, column=8, value=element_takecoeff[1])
                ws.cell(row=row, column=10, value=shotname_after)
                ws.cell(row=row, column=14, value=bet_size)
                ws.cell(row=row, column=16, value=(float(bankrol_first) - bet_size))
            else:
                ws.cell(row=row, column=9, value=element_takecoeff[1])
                ws.cell(row=row, column=11, value=shotname_after)
                ws.cell(row=row, column=17, value=bet_size)
                ws.cell(row=row, column=19, value=(float(bankrol_second)-bet_size))
        elif element_takecoeff[0] == '1XBet':
            if '1xlite' in event[6]:
                ws.cell(row=row, column=8, value=element_takecoeff[1])
                ws.cell(row=row, column=10, value=shotname_after)
                ws.cell(row=row, column=14, value=bet_size)
                ws.cell(row=row, column=16, value=(float(bankrol_first) - bet_size))
            else:
                ws.cell(row=row, column=9, value=element_takecoeff[1])
                ws.cell(row=row, column=11, value=shotname_after)
                ws.cell(row=row, column=17, value=bet_size)
                ws.cell(row=row, column=19, value=(float(bankrol_second)-bet_size))
        elif element_takecoeff[0] == 'BETBOOM':
            if 'betboom' in event[6]:
                ws.cell(row=row, column=8, value=element_takecoeff[1])
                ws.cell(row=row, column=10, value=shotname_after)
                ws.cell(row=row, column=14, value=bet_size)
                ws.cell(row=row, column=16, value=(float(bankrol_first) - bet_size))
            else:
                ws.cell(row=row, column=9, value=element_takecoeff[1])
                ws.cell(row=row, column=11, value=shotname_after)
                ws.cell(row=row, column=17, value=bet_size)
                ws.cell(row=row, column=19, value=(float(bankrol_second)-bet_size))
        for i in range(len(event_data)-2,len(event_data)):
            ws.cell(row=row, column=(i+6), value=event_data[i])
        wb.save(f"{path}")
        asyncio.run(event_result(urls=event, row=row, bmaker=element_takecoeff[0]))