import os
import logging
from time import sleep

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill

logger = logging.getLogger('Client UI.components.statistic_management.statistic')


class StatisticManager:
    """Формирование статистики и запись в файл xlsx"""
    columns = ("№  ", "Команды", "Дата", "ТМ/Коэфф.", "ТБ/Коэфф.", "Скриншот", "Ссылка", "Ставка", "Баланс до",
               "Баланс после", "Время размещения", "Сумма голов", "Результат", "Баланс")
    font = 'Calibri'
    font_size = 12
    # задаем параметры границ ячеек
    thins = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    def insert_data(self, event_data: list):
        """Заполнение данными документа xlsx"""
        if not os.path.exists(settings.STATS_DIR):
            os.mkdir(settings.STATS_DIR)

        if not os.path.exists(settings.STATS_FILE_NAME):
            wb = Workbook()
            ws_betting = wb.active
            ws_betting.title = "Ставки на деньги"
            for col in range(1, len(self.columns)+1):
                cell = ws_betting.cell(row=1, column=col, value=self.columns[col-1])

                # выравнивание по ширине ячейки
                cell.font = Font(name=self.font, size=self.font_size)
                len_cell = len(str(cell.value))
                new_width_col = len_cell * self.font_size ** (self.font_size * 0.01)
                ws_betting.column_dimensions[cell.column_letter].width = new_width_col
                ws_betting.column_dimensions[cell.column_letter].alignment = Alignment(horizontal="center")

            ws_betting.column_dimensions['B'].width = 30


            ws_imitation = wb.copy_worksheet(ws_betting)
            ws_imitation.title = "Имитация ставок"

        else:
            for _ in range(5):
                try:
                    wb = load_workbook(settings.STATS_FILE_NAME)
                    break
                except BaseException as ex:
                    logger.info(f'Ошибка открытия файла статистики, {ex}')
                    if _ == 4:
                        logger.info(f'Не удалось получить доступ к файлу статистики, {ex}')
                        return
                    sleep(0.5)

        if event_data[0]['bet_imitation']:
            ws = wb["Имитация ставок"]
        else:
            ws = wb["Ставки на деньги"]

        event_nums = []
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=len(ws['A'])):
            for cell in row:
                if cell.value:
                    event_nums.append(int(cell.value))
        if event_nums:
            event_num = max(event_nums) + 1
        else:
            event_num = 1

        for data in event_data:
            empty_row_num = len(ws['A']) + 1
            ws.cell(row=empty_row_num, column=1).value = event_num
            ws.cell(row=empty_row_num, column=2).value = data['teams']
            ws.cell(row=empty_row_num, column=3).value = data['date']
            if data['total_koeff_type'] == 'under':
                ws.cell(row=empty_row_num, column=4).value = data['total_koeff']
                ws.cell(row=empty_row_num, column=5).value = '-'
            else:
                ws.cell(row=empty_row_num, column=4).value = '-'
                ws.cell(row=empty_row_num, column=5).value = data['total_koeff']
            ws.cell(row=empty_row_num, column=6).value = data['screenshot_name']
            ws.cell(row=empty_row_num, column=7).value = data['url']
            ws.cell(row=empty_row_num, column=8).value = data['bet_size']
            ws.cell(row=empty_row_num, column=9).value = data['start_balance']
            ws.cell(row=empty_row_num, column=10).value = data['balance_after_bet']
            ws.cell(row=empty_row_num, column=11).value = data['betting_time']

        empty_row_num = len(ws['A']) + 1
        for col in range(1, len(self.columns)+1):
            cell = ws.cell(row=empty_row_num, column=col)
            cell.fill = PatternFill(fgColor="000000", fill_type="solid")
            cell.border = Border(top=self.double, bottom=self.double, left=self.thins, right=self.thins)

        wb.save(settings.STATS_FILE_NAME)


if __name__ == '__main__':
    from pathlib import Path
    BASE_DIR = Path(__file__).resolve().parent.parent
    class settings:
        STATS_DIR = BASE_DIR / "statistic"
        STATS_FILE_NAME = STATS_DIR / "statistic.xlsx"
    event_data = [{'balance_after_bet': 90,
                   'betting_time': '100',
                   'screenshot_name': 'D:/screenshots/leon.png',
                   'bet_imitation': True,
                   'start_balance': 100,
                   'total_koeff': 2.1,
                   'total_koeff_type': 'under',
                   'bet_size': 10,
                   'total_nominal': 2.5,
                   'url': 'link',
                   'teams': 'ЦСКА - Спартак',
                   'bookmaker': 'leon',
                   'date': '2025-01-12'},
                  {'balance_after_bet': 110,
                   'betting_time': '80',
                   'screenshot_name': 'D:/screenshots/olimp.png',
                   'bet_imitation': True,
                   'start_balance': 120,
                   'total_koeff': 2.1,
                   'total_koeff_type': 'over',
                   'bet_size': 10,
                   'total_nominal': 2.5,
                   'url': 'link',
                   'teams': 'ЦСКА - Спартак',
                   'bookmaker': 'olimp',
                   'date': '2025-01-12'}]
    statistic = StatisticManager()
    statistic.insert_data(event_data)
else:
    from .. import settings




# ("№  ", "Команды БК1", "Команды БК2", "ТМ/Коэфф. БК1", "ТБ/Коэфф. БК2", "ТБ/Коэфф. БК1", "ТМ/Коэфф. БК2",
#                "Коэффициент БК1", "Коэффициент БК2", "Скриншот БК1", "Скриншот БК2", "Ссылка БК1", "Ссылка БК2",
#                "Ставка БК1", "Результат БК1", "Банкрол БК1", "Ставка БК2", "Результат БК2", "Сумма голов", "Дата БК1",
#                "Дата БК2")